"""

Explains how to edit, read ,write and do operations on xls sheets

"""

import openpyxl
import os

workbook=openpyxl.load_workbook("work_book location")
type(workbook)

sheet=workbook.get_sheet_by_name("sheet_name")

# if we don;t know sheet names
workbook.get_sheet_names()  # gives list of sheet names

cell=sheet['A1']  # to get the data from cells
cell.value

sheet['B1'].value
sheet.cell(row=1,column=2) # returns cell object 
# another way of accessing the cells

for i in range(1,5):
  print(i,sheet.cell(row=i,column=2).value)
  
  
  
# write opeartions on xsl sheet

wb=openpyxl.Workbook()
wb.get_sheet_names()
sheet=wb.get_sheet_by_name('sheet_name')

sheet['A1']='hello'

wb.save('file_name.xlsx') # to save the changes to workbook

sheet2=wb.create_sheet()  # creating new sheet
# another way
wb.create_sheet(index=0,title="name") # we can change the position of sheets using index and change the title as below

sheet2.title="new_sheet_name"  # changing sheet name

